import openpyxl

wb = openpyxl.load_workbook("data/pricing/online_retail_II.xlsx", read_only=True)
sheet = wb.active

headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
first_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

print("=== COLUMNS ===")
for col in headers:
    print(f"  {col}")

print("\n=== FIRST FULL ROW ===")
for col, val in zip(headers, first_row):
    print(f"  {col}: {val}")
        